import pandas as pd
import re
import chardet
from openpyxl import load_workbook
import unicodedata

class ShipmentDirection:
    def __init__(self):
        self.shipments = {}
        self.consolidated_shipment_orders = {}

    def add_shipment(self, picknum, total_orders, tracking_needed_orders=None, special_instructions=None):
        self.shipments[picknum] = {
            'total_orders': total_orders,
            'tracking_needed_orders': tracking_needed_orders,
            'special_instructions': special_instructions
        }
    def update_tracking_needed_orders(self,picknum,tracking_needed_orders):
        self.shipments[picknum]['tracking_needed_orders'] = tracking_needed_orders

    def update_special_instructions(self,picknum,special_instructions):
        self.shipments[picknum]['special_instructions'] = special_instructions

    # consolidated_shipname_ordernum 为一个字典
    def add_consolidated_shipment_orders(self, consolidated_shipname_ordernum):
        self.consolidated_shipment_orders.update(consolidated_shipname_ordernum)

    def get_shipment_by_picktime(self, picknum):
        return self.shipments[picknum]

    def get_all_shipments(self):
        return self.shipments

    def get_consolidated_shipment_orders(self):
        return self.consolidated_shipment_orders

def detect_encoding(file):
    with open(file, 'rb') as f:
        result = chardet.detect(f.read())
    return result['encoding']

def pickingcsv_loading(csv_list, header_list):
    dataframes = []

    for i, file in enumerate(csv_list):
        print(f"Processing file: {file}")
        encoding = detect_encoding(file)
        print(f"Detected encoding: {encoding}")
        try:
            df = pd.read_csv(file, encoding=encoding, names=header_list, dtype="str")
            df['SourceFile'] = str(i + 1)  # 增加一列并赋值，从1开始
            dataframes.append(df)
        except Exception as e:
            print(f"Error reading file {file}: {e}")
            continue
    
    merged_df = pd.concat(dataframes, ignore_index=True)
    return merged_df

def remove_after_last_digit(text):
    # 使用正则表达式进行匹配
    return re.sub(r'(\d)[^\d]*$', r'\1', text)

def pivot_for_manual(pickdf):
    pivotdf = pickdf[['OSHNA1','OSHNA2','OSHAD1','OSHAD2','SourceFile','OSONO']]
    pivotdf = pivotdf.groupby(['SourceFile', 'OSONO']).first().reset_index()
    pivotdf['SHIPTO名前'] = pivotdf['OSHNA1'] + pivotdf['OSHNA2'].fillna("")
    pivotdf['SHIPTO住所'] = pivotdf['OSHAD1'] + pivotdf['OSHAD2'].fillna("") 
    pivotdf = pivotdf.drop(columns=['OSHNA1','OSHNA2','OSHAD1','OSHAD2'])
    pivotdf = pivotdf.rename(columns={'SourceFile':'PICK回目','OSONO':'#ORD'})
    return pivotdf

def order_num_count(pickdf):
    order_countdf = pickdf.groupby('SourceFile')['OSONO'].nunique().reset_index()
    
    order_count = order_countdf.set_index('SourceFile')['OSONO'].to_dict()
    return order_count

# Set OMEMO3,4 and 3rd line of SHIPTO as special note
def special_note(pickdf):
    special_note_df = pickdf[['OSONO','OSHAD3','OMEMO1','OMEMO2','OMEMO3','OMEMO4','SourceFile']].drop_duplicates()

    special_note_df['PROFILE'] = special_note_df['OSHAD3'].fillna('').apply(lambda x: ' '.join(re.findall(r'\*(.*?)\*', x)))

    special_note_df['SpecialNote'] = special_note_df['OMEMO1'].fillna('').apply(lambda x: unicodedata.normalize('NFKC',x)) + special_note_df['OMEMO2'].fillna('').apply(lambda x: unicodedata.normalize('NFKC',x)) + special_note_df['OMEMO3'].fillna('').apply(lambda x: unicodedata.normalize('NFKC',x)) + special_note_df['OMEMO4'].fillna('').apply(lambda x: unicodedata.normalize('NFKC',x))

    special_note_df['SpecialNote'] = special_note_df['SpecialNote'].apply(lambda x: ' '.join(re.findall(r'\*(.*?)\*', x)))
    special_note_df['SpecialNote'] = special_note_df['PROFILE'].fillna('') + special_note_df['SpecialNote'].fillna('')

    special_note_df['SpecialNote'] = special_note_df['SpecialNote'].str.replace("送り状要","")
    special_note_df['SpecialNote'] = special_note_df['SpecialNote'].str.replace("同梱不可","")
    special_note_df['SpecialNote'] = special_note_df['SpecialNote'].str.strip()
    special_note_df = special_note_df[special_note_df['SpecialNote'] != '']

    special_note_df['SpecialNote'] = special_note_df['OSONO'] + ":" + special_note_df['SpecialNote']
    
    special_note_df = special_note_df.groupby('SourceFile')['SpecialNote'].apply(list)
    special_note_dict = special_note_df.to_dict()
    return special_note_dict

# Set OMEMO3 as waybill request
def waybill_request(pickdf):
    waybill_request_df = pickdf[['OSONO','OSHAD3','OMEMO1','OMEMO2','OMEMO3','OMEMO4','SourceFile']].drop_duplicates()
    #waybill_request_df = waybill_request_df.dropna(subset = ['OMEMO3'])
    waybill_request_df['WB'] = waybill_request_df['OSHAD3'].fillna('') + waybill_request_df['OMEMO1'].fillna('')+ waybill_request_df['OMEMO2'].fillna('') + waybill_request_df['OMEMO3'].fillna('') + waybill_request_df['OMEMO4'].fillna('')
    waybill_request_df = waybill_request_df[waybill_request_df['WB'].str.contains("送り状要")]
    waybill_request_df = waybill_request_df.groupby('SourceFile')['OSONO']
    waybill_request_dict = waybill_request_df.apply(list).to_dict()
    return waybill_request_dict

def consolidate_shipment(pickdf,name_block_list=None,address_block_list=None,fixed_consolidate_dict=None):
    name_block_pattern = '|'.join(name_block_list)
    address_block_pattern = '|'.join(address_block_list)
    consolidate_df = pickdf[~pickdf['OSHNA1'].str.contains(name_block_pattern)]
    consolidate_df = consolidate_df[~consolidate_df['OSHAD1'].str.contains(address_block_pattern)]
    consolidate_df = consolidate_df[['OSONO','OTELNO','OSHNA1','OSHAD1','OSHAD2','OSHAD3','OMEMO1','OMEMO2','OMEMO3','OMEMO4','SourceFile']].drop_duplicates()

    specific_note = "同梱不可"

    contains_note = consolidate_df.applymap(lambda x: specific_note in x if isinstance(x, str) else False)
    rows_to_drop = contains_note.any(axis=1)
    consolidate_df = consolidate_df.drop(index=consolidate_df[rows_to_drop].index)

    consolidate_df['OTELNO'] = consolidate_df['OTELNO'].str.strip()
    consolidate_df['OTELNO'] = consolidate_df['OTELNO'].str.replace("-","")
    fixed_consolidate_df = consolidate_df[['OSONO','OTELNO']]
    consolidate_order_dict ={}
    for location, phone in fixed_consolidate_dict.items():
        consolidate_order_dict[location] = fixed_consolidate_df[fixed_consolidate_df['OTELNO'] == phone]['OSONO'].to_list()

    
    other_df = consolidate_df[~consolidate_df['OTELNO'].isin(list(fixed_consolidate_dict.values()))]
    other_df.loc[:, 'SHIPADD'] = other_df['OSHAD1'] + other_df['OSHAD2'].fillna('')
    other_df.loc[:, 'SHIPADD'] = other_df.loc[:, 'SHIPADD'].apply(lambda x: unicodedata.normalize('NFKC',x))
    other_df['SHIPADD'] = other_df['SHIPADD'].str.replace(" ","")
    other_df['SHIPADD'] = other_df['SHIPADD'].apply(remove_after_last_digit)

    other_df['OSHNA1'] = other_df['OSHNA1'].apply(lambda x: unicodedata.normalize('NFKC',x))
    other_df = other_df[other_df.duplicated(subset=['SHIPADD'],keep=False)]

    if other_df.empty is False:
        other_dict = other_df.groupby('SHIPADD').apply(lambda x: {'OSHNA1': x['OSHNA1'].iloc[0], 'OSONO': x['OSONO'].tolist()}).to_dict()
        other_dict = {v['OSHNA1']: v['OSONO'] for v in other_dict.values()}
        consolidate_order_dict.update(other_dict)
    return consolidate_order_dict

def template_fulfillment(excel_template,shipment_direction,pivotdf,outputpath):
    shipment_all=shipment_direction.get_all_shipments()
    wb = load_workbook(filename=excel_template)
    ws = wb['Epiroc PickList送付']

    for i in range(1,len(shipment_all)+1,1):
        j = str(i)
        shipment_direction_dict_perpicktime = shipment_direction.get_shipment_by_picktime(j)
        #print(shipment_direction_dict_perpicktime)
        # 件数
        total_orders = shipment_direction_dict_perpicktime['total_orders']
        ws.cell(row=(2*i+9),column=2,value=(j+'回目'))
        ws.cell(row=(2*i+9),column=4,value=total_orders)
        # Special Instructions
        special_instruction = shipment_direction_dict_perpicktime["special_instructions"]
        if special_instruction is not None:
            for index,item in enumerate(special_instruction,start=1):
                ws.cell(row=(21+index),column=(2),value=(j+'回目'))
                ws.cell(row=(21+index),column=(3),value=(item))
        # Waybill request
        waybill_request_list = shipment_direction_dict_perpicktime["tracking_needed_orders"]
        start_row = 36
        start_col = 2+3*(i-1)
        if waybill_request_list is not None:
            index = 0
            for i in range(start_row, start_row + 3):
                for j in range(start_col, start_col + 3):
                    if index < len(waybill_request_list):
                        ws.cell(row=i, column=j, value=waybill_request_list[index])
                        index += 1
    # Consolidated shipment
    consolidated_shipment_dict = shipment_direction.get_consolidated_shipment_orders()
    # 定义起始行和列
    start_row = 45
    key_col = 3  # C 列
    value_col = 4  # D 列

    # 将字典内容写入指定区域
    for index, (key, values) in enumerate(consolidated_shipment_dict.items(), start=start_row):
        ws.cell(row=index, column=key_col, value=key)
        ws.cell(row=index, column=value_col, value=', '.join(values))

    pivot_sheet_name = 'pivotcsv'
    if pivot_sheet_name in wb.sheetnames:
        ws_pivot = wb[pivot_sheet_name]
    else:
        ws_pivot = wb.create_sheet(title=pivot_sheet_name)

    for col_idx, col_name in enumerate(pivotdf.columns, start=1):
        ws_pivot.cell(row=1, column=col_idx, value=col_name)
    
    for r_idx, row in enumerate(pivotdf.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws_pivot.cell(row=r_idx, column=c_idx, value=value)

    ws_pivot.column_dimensions['C'].width = 56
    ws_pivot.column_dimensions['D'].width = 56
        
    wb.save(outputpath)

if __name__ == '__main__':
    csv0 = r"\\ssisjpfs0004\JPN\MRBA\Logistics\CMT Logistics\FromBPCS\DOWNLOADS\Yamato\送信済みデータ\lypl20240726 1100.csv"
    csv1 = r"\\ssisjpfs0004\JPN\MRBA\Logistics\CMT Logistics\FromBPCS\DOWNLOADS\Yamato\送信済みデータ\lypl20240726 1430.csv"
    csv2 = r"\\ssisjpfs0004\JPN\MRBA\Logistics\CMT Logistics\FromBPCS\DOWNLOADS\Yamato\送信済みデータ\lypl20240726 1530.csv"
    #csv3 = r"\\ssisjpfs0004\JPN\MRBA\Logistics\CMT Logistics\FromBPCS\DOWNLOADS\Yamato\送信済みデータ\lypl20240723 1530.csv"

    tempcsv = "pick0726.csv"
    excel_template = r"C:\Users\jpeqz\OneDrive - Epiroc\Python\Outbounddoc\送り状鑑(更新版_py).xlsx"
    output_path = r"C:\Users\jpeqz\OneDrive - Epiroc\Python\Outbounddoc\output.xlsx"
    header_list = ["OSONO","OSHIP","OTYPE","OCUSPO","OCUSNO","OCUSNA","OCUSA1","OCUSA2","ODATE","OSHNA1","OSHNA2","OSHZIP","OSHAD1","OSHAD2","OSHAD3","OSHATN","OTELNO","OSOLNE","OITMN","OSERN","OLOCN","OIDESC","OQTY","ODDATE","ODTIME","OTRNSP","OMEMO1","OMEMO2","OMEMO3","OMEMO4","OSHIPR","OPGC","OPLC"]
    name_block_list = ["戸髙","鳥形","峩朗"]
    address_block_list = ["高知県吾川郡仁淀川町","峩朗"]
    fixed_consolidate_dict = {"ｴﾋﾟﾛｯｸ福岡":"0925580621", "ｴﾋﾟﾛｯｸ大阪":"0727754511","ｴﾋﾟﾛｯｸ仙台": "0223473755", "DRM兵庫":"0795360461","虎乃門千葉": "0436222141"}
    csv_list = [csv0,csv1,csv2]
    
    shipment_direction = ShipmentDirection()
    pickdf = pickingcsv_loading(csv_list,header_list)
    pivotdf = pivot_for_manual(pickdf)
    
    pickdf.to_csv(tempcsv,encoding='UTF_8_sig',index=False)
    pickdf = pd.read_csv(tempcsv,encoding='UTF_8_sig',dtype="str")
    ordercount = order_num_count(pickdf)
    for picktime in ordercount:
        shipment_direction.add_shipment(picktime,ordercount[picktime])
    
    waybill_request_dict = waybill_request(pickdf)
    for picktime in waybill_request_dict:
        shipment_direction.update_tracking_needed_orders(picktime,waybill_request_dict[picktime])
    
    special_note_dict = special_note(pickdf)
    for picktime in special_note_dict:
        shipment_direction.update_special_instructions(picktime,special_note_dict[picktime])

    #print(shipment_direction.get_shipment_by_picktime('1'))
    consolidate_shipment_dict = consolidate_shipment(pickdf,name_block_list,address_block_list,fixed_consolidate_dict)
    shipment_direction.add_consolidated_shipment_orders(consolidate_shipment_dict)
    print(shipment_direction.get_all_shipments())
    print(shipment_direction.get_consolidated_shipment_orders())
    template_fulfillment(excel_template,shipment_direction,pivotdf,output_path)

